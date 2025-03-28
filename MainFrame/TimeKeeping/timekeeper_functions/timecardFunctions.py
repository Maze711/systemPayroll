from MainFrame.Resources.lib import *
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from MainFrame.notificationMaker import notificationLoader
from MainFrame.Database_Connection.DBConnection import create_connection
from MainFrame.TimeKeeping.schedValidator.checkSched import chkSched
from MainFrame.TimeKeeping.schedValidator.schedule import schedule
from MainFrame.TimeKeeping.timeCardMaker.filter import FilterDialog
from MainFrame.TimeKeeping.timeSheet.timeSheet import TimeSheet
from MainFrame.systemFunctions import timekeepingFunction, globalFunction
from MainFrame.TimeKeeping.timekeeper_functions.processTimeSheetLoader import processTimeSheetLoader

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class populateList:
    def __init__(self, parent):
        self.parent = parent
        self.data_cache = {}
        self.cost_center_cache = None

    def import_dat_file(self, dialog):
        fileName, _ = QFileDialog.getOpenFileName(dialog, "Open DAT File", "", "DAT Files (*.DAT)")
        try:
            if fileName:
                # Block signals before showing notification dialog
                self.parent.dateFromCC.blockSignals(True)
                self.parent.dateToCC.blockSignals(True)

                notification_dialog = notificationLoader(fileName)
                notification_dialog.importSuccessful.connect(self.update_after_import)
                notification_dialog.finished.connect(lambda: [
                    self.parent.dateFromCC.blockSignals(False),
                    self.parent.dateToCC.blockSignals(False)
                ])
                notification_dialog.exec_()
            else:
                QMessageBox.information(dialog, "No File Selected", "Please select a DAT file to import.")
                return
        except Exception as e:
            print(f"error to ng import {e}")
            # Ensure signals are unblocked even if error occurs
            self.parent.dateFromCC.blockSignals(False)
            self.parent.dateToCC.blockSignals(False)

    def update_after_import(self):
        self.populate_year_combo_box()
        self.populate_date_combo_boxes()

    def populate_year_combo_box(self):
        connection = create_connection('NTP_LOG_IMPORTS')
        if not connection:
            return

        try:
            cursor = connection.cursor()
            cursor.execute("SHOW TABLES LIKE 'table_%'")
            tables = cursor.fetchall()

            year_months = set()
            for (table_name,) in tables:
                match = re.search(r'table_(\d{4})_(\d{2})', table_name)
                if match:
                    year_months.add(f"{match.group(1)}_{match.group(2)}")

            sorted_year_months = sorted(year_months, reverse=True)

            if set(self.parent.yearCC.itemText(i) for i in range(self.parent.yearCC.count())) != set(
                    sorted_year_months):
                self.parent.yearCC.clear()
                self.parent.yearCC.addItems(sorted_year_months)

            self.parent.yearCC.setCurrentText("")

        except Exception as e:
            logging.error(f"Error populating year combo box: {e}")
            QMessageBox.critical(self.parent, "Error", f"Failed to populate year combo box: {str(e)}")

        finally:
            cursor.close()
            connection.close()

    def populate_date_combo_boxes(self):
        selected_year_month = self.parent.yearCC.currentText()
        if not selected_year_month:
            return

        if selected_year_month in self.data_cache:
            days = self.data_cache[selected_year_month]
        else:
            connection = create_connection('NTP_LOG_IMPORTS')
            if not connection:
                return

            try:
                cursor = connection.cursor()
                table_name = f"table_{selected_year_month}"
                cursor.execute(f"SELECT DISTINCT DATE_FORMAT(date, '%d') AS day FROM {table_name} ORDER BY day ASC")
                days = [day[0] for day in cursor.fetchall()]
                self.data_cache[selected_year_month] = days
            except Exception as e:
                logging.error(f"Error populating date combo boxes for {selected_year_month}: {e}")
                QMessageBox.critical(self.parent, "Error", f"Failed to populate date combo boxes: {str(e)}")
                return
            finally:
                cursor.close()
                connection.close()

        current_from_selection = self.parent.dateFromCC.currentText()
        current_to_selection = self.parent.dateToCC.currentText()

        if set(self.parent.dateFromCC.itemText(i) for i in range(self.parent.dateFromCC.count())) != set(days):
            self.parent.dateFromCC.clear()
            self.parent.dateFromCC.addItems(sorted(days))

        if set(self.parent.dateToCC.itemText(i) for i in range(self.parent.dateToCC.count())) != set(days):
            self.parent.dateToCC.clear()
            self.parent.dateToCC.addItems(sorted(days))

        self.parent.dateFromCC.setCurrentText("")
        self.parent.dateToCC.setCurrentText("")


    def populateCostCenterBox(self):
        """Populate the costCenterBox with values from the dept_name column in the emp_posnsched table."""
        if self.cost_center_cache is not None:
            # Use cached data if available
            self.parent.costCenterBox.clear()
            self.parent.costCenterBox.addItems(self.cost_center_cache)
            self.parent.costCenterBox.setCurrentIndex(-1)
            logging.info("Cost center box populated from cache.")
            return

        connection = create_connection('NTP_EMP_LIST')
        if not connection:
            logging.error("Error: Unable to connect to FILE201 database.")
            return

        try:
            cursor = connection.cursor()

            # Optimized query to fetch distinct dept_name values
            query = "SELECT DISTINCT dept_name FROM emp_posnsched ORDER BY dept_name"
            cursor.execute(query)

            # Fetch all results at once
            dept_names = [row[0] for row in cursor.fetchall()]

            # Cache the results
            self.cost_center_cache = dept_names

            # Clear the current items in the QComboBox
            self.parent.costCenterBox.clear()
            self.parent.costCenterBox.addItems(dept_names)
            self.parent.costCenterBox.setCurrentIndex(-1)

        except Exception as e:
            QMessageBox.critical(self.parent, "Error",
                                 f"An error occurred while populating the cost center box: {e}. "
                                 "Please try again or contact the system administrator.")
        finally:
            cursor.close()
            connection.close()

    def clear_cache(self):
        """Clear all cached data"""
        self.data_cache.clear()
        self.cost_center_cache = None
        logging.info("All caches cleared.")

    def populate_table_loader(self):
        if hasattr(self, 'TablePopulationLoader') and self.TablePopulationLoader.isVisible():
            return

        self.TablePopulationLoader = TablePopulationLoader(self.parent.original_data, self.parent)
        self.TablePopulationLoader.show()

    def populate_table_with_data(self, data):
        try:
            logging.info("Populating table with data.")
            self.parent.TimeListTable.setUpdatesEnabled(False)
            self.parent.TimeListTable.setSortingEnabled(False)

            # Sort data alphabetically by employee name (assuming it's the second column)
            sorted_data = sorted(data, key=operator.itemgetter(2))

            self.parent.TimeListTable.setRowCount(len(sorted_data))

            # Prepopulate combo box options with hour-only format
            combo_items = [f"{hour:02d}:00:00" for hour in range(24)]

            # Set the custom delegate for the sched_in and sched_out columns (7 and 8)
            combo_delegate = ComboBoxDelegate(combo_items, self.parent.TimeListTable)

            # connects the cbValueChanged signal to slot
            combo_delegate.cbValueChanged.connect(self.updateSchedInAndOut)

            # Setting delegate outside the loop
            self.parent.TimeListTable.setItemDelegateForColumn(7, combo_delegate)
            self.parent.TimeListTable.setItemDelegateForColumn(8, combo_delegate)

            # Populate table data
            for row_position, row_data in enumerate(sorted_data):
                for col, value in enumerate(row_data):
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)

                    # Allow editing only for columns 6 and 7
                    if col == 7 or col == 8:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                    else:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Disable editing for other columns

                    self.parent.TimeListTable.setItem(row_position, col, item)

            header = self.parent.TimeListTable.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Fixed)
            header.setSectionsClickable(False)

            logging.info(f"Table populated with {len(sorted_data)} rows.")
        except Exception as e:
            logging.error(f"Error populating table with data: {str(e)}")
            QMessageBox.critical(self.parent, "Error", f"An error occurred while populating the table: {str(e)}")
        finally:
            self.parent.TimeListTable.setSortingEnabled(True)
            self.parent.TimeListTable.setUpdatesEnabled(True)

        # Ensure the edit triggers are set to make cells easier to edit
        self.parent.TimeListTable.setEditTriggers(QAbstractItemView.AllEditTriggers)

    def updateSchedInAndOut(self, row, updated_value):
        """ Updates the Sched in/out once the combo box has been changed"""
        selected_year_month = self.parent.yearCC.currentText()
        id = int(self.parent.TimeListTable.item(row, 0).text())
        sched_in = self.parent.TimeListTable.item(row, 7).text()
        sched_out = self.parent.TimeListTable.item(row, 8).text()

        # Use create_connection to establish a connection
        connection = create_connection('NTP_LOG_IMPORTS')

        if not connection:
            logging.error("Error: Unable to connect to NTP_LOG_IMPORTS database.")
            return

        try:
            cursor = connection.cursor()

            table_name = f"table_{selected_year_month}"
            update_sched_in_and_out = f"UPDATE {table_name} SET sched_in = %s, sched_out = %s WHERE ID = %s"
            cursor.execute(update_sched_in_and_out, (sched_in, sched_out, id))
            connection.commit()

            if cursor.rowcount > 0:
                print(f'Updated Successfully sched in and out for id: {id}')

        except Exception as e:
            print(f"Error updating sched in and out with for id: {id}: {e}")
            QMessageBox.critical(self.parent, "Error",
                                 f"Error updating sched in/out with for 8id: {id}: {e}")
            return

        finally:
            # Ensure the connection is closed after use
            if connection:
                connection.close()


class ComboBoxDelegate(QStyledItemDelegate):
    cbValueChanged = pyqtSignal(int, str) # signal to emit the row and updated value
    def __init__(self, items, parent=None):
        super(ComboBoxDelegate, self).__init__(parent)
        self.items = items

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self.items)
        combo.currentIndexChanged.connect(lambda: self.commitData.emit(combo))
        return combo

    def setEditorData(self, editor, index):
        value = index.data()
        editor.setCurrentText(value)

    def setModelData(self, editor, model, index):
        current_value = index.data()
        new_value = editor.currentText()

        # Only emit signal if the value is actually changed
        if current_value != new_value:
            model.setData(index, new_value)
            self.cbValueChanged.emit(index.row(), new_value)


class buttonTimecardFunction:
    def __init__(self, parent):
        self.parent = parent  # Reference to the main UI class
        self.time_computation = TimeComputation(parent)

    def updateSchedule(self):
        # Access UI elements through self.parent
        selectedCostCenterFrom = self.parent.costCenterFrom.currentText()
        selectedCostCenterBox = self.parent.costCenterBox.currentText()
        selectedCostCenterTo = self.parent.costCenterTo.currentText()

        if not selectedCostCenterFrom or not selectedCostCenterBox or not selectedCostCenterTo:
            print("Please select all required fields.")
            return

        # Print the selected data
        print(f"Selected Cost Center From: {selectedCostCenterFrom}")
        print(f"Selected Cost Center To: {selectedCostCenterTo}")
        print(f"Selected Cost Center Box: {selectedCostCenterBox}")

        # Show caution message box
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Caution")
        msg.setText(f"CAUTION! It will change the schedule of {selectedCostCenterBox}. Do you want to continue?")
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        result = msg.exec_()

        # Proceed only if the user clicks "Yes"
        if result == QMessageBox.Yes:
            # Database connection (ensure create_connection is defined/imported)
            connection = create_connection('NTP_EMP_LIST')
            if not connection:
                logging.error("Error: Unable to connect to FILE201 database.")
                return

            try:
                cursor = connection.cursor()

                # Query to fetch existing schedule
                query = """
                    SELECT sched_in, sched_out
                    FROM emp_posnsched
                    WHERE dept_name = %s
                """
                cursor.execute(query, (selectedCostCenterBox,))
                result = cursor.fetchone()

                if result:
                    sched_in, sched_out = result
                    print(f"Existing Scheduled In: {sched_in}")
                    print(f"Existing Scheduled Out: {sched_out}")

                    # Update sched_in and sched_out based on selectedCostCenterFrom and selectedCostCenterTo
                    update_query = """
                        UPDATE emp_posnsched
                        SET sched_in = %s, sched_out = %s
                        WHERE dept_name = %s
                    """
                    cursor.execute(update_query, (selectedCostCenterFrom, selectedCostCenterTo, selectedCostCenterBox))
                    connection.commit()

                    print(f"Updated Scheduled In to: {selectedCostCenterFrom}")
                    print(f"Updated Scheduled Out to: {selectedCostCenterTo}")
                else:
                    print("No schedule found for the selected Cost Center Box.")

            except Exception as e:
                QMessageBox.critical(self.parent, "Update Error",
                                     f"An error occurred while updating the schedule: {e}. "
                                     "Please try again or contact the system administrator.")
            finally:
                connection.close()
        else:
            print("Update canceled by user.")

    def export_to_excel(self):
        try:
            # Create an Excel workbook and sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Timecard Data"

            # Get year, month, start and end day from the UI
            year, month = self.parent.yearCC.currentText().split('_')
            start_day = self.parent.dateFromCC.currentText()
            end_day = self.parent.dateToCC.currentText()

            # Define headers for employee data
            headers = ["No.", "Names", "Employee No", "BioNum", "Machcode", "Check_In",
                       "Check_Out"]
            start_date = datetime(int(year), int(month), int(start_day))
            end_date = datetime(int(year), int(month), int(end_day))

            # Create date range
            date_range = pd.date_range(start_date, end_date)

            # Prepare the date headers and day name headers
            date_headers = []
            day_name_headers = []

            for date in date_range:
                formatted_date = date.strftime("%d-%b")  # Format as "01-Aug"
                day_name = calendar.day_name[date.weekday()][:3].upper()  # Get abbreviated day name (e.g., "THU")
                date_headers.append(formatted_date)
                day_name_headers.append(day_name)

            # Write headers to the first row of the Excel sheet
            ws.append(headers + date_headers)  # Write the static headers for employee data followed by date headers

            # Write the day name headers on the second row, aligned directly under each date
            ws.append([''] * len(headers) + day_name_headers)  # Fill with empty strings for the first part

            # Initialize employee data dictionary
            employee_data = {}

            # Get the number of rows from the table
            rows = self.parent.TimeListTable.rowCount()

            # Loop through the table to collect data per employee
            for row in range(rows):
                # Access each row from the TimeListTable
                bio_num = self.parent.TimeListTable.item(row, 1).text()  # BioNum
                emp_name = self.parent.TimeListTable.item(row, 2).text()  # Employee name
                trans_date = pd.to_datetime(self.parent.TimeListTable.item(row, 3).text())  # Date
                machcode = self.parent.TimeListTable.item(row, 4).text()  # Machcode
                check_in = self.parent.TimeListTable.item(row, 5).text()  # Check-in time
                check_out = self.parent.TimeListTable.item(row, 6).text()  # Check-out time
                sched_in = self.parent.TimeListTable.item(row, 7).text()  # Time-In
                sched_out = self.parent.TimeListTable.item(row, 8).text()  # Time-Out

                # Initialize employee entry if not already present
                if bio_num not in employee_data:
                    employee_data[bio_num] = {
                        "No": len(employee_data) + 1,
                        "Names": emp_name,
                        "Employee No": bio_num,
                        "BioNum": bio_num,
                        "Machcode": machcode,
                        "Check_In": check_in,
                        "Check_Out": check_out,
                        "Sched-In": sched_in,
                        "Sched-Out": sched_out,
                        "Dates": {str(date.date()): "" for date in date_range}  # Empty slots for each date
                    }

                # Populate check-in and check-out for the corresponding date
                if check_in and check_out:
                    # Ensure the date exists in the date range
                    if str(trans_date.date()) in employee_data[bio_num]["Dates"]:
                        employee_data[bio_num]["Dates"][str(trans_date.date())] = f"{check_in} - {check_out}"

            # Write employee data to the Excel sheet
            for emp in employee_data.values():
                row_data = [
                    emp["No"], emp["Names"], emp["Employee No"], emp["BioNum"],
                    emp["Machcode"], emp["Check_In"], emp["Check_Out"]
                ]

                # Append check-in and check-out times for each date
                for date in date_range:
                    row_data.append(emp["Dates"].get(str(date.date()), ""))  # Use date.date() to match keys

                ws.append(row_data)  # Write the data row to the Excel sheet

            # Apply text alignment (center) to all cells
            for row in ws.iter_rows(min_row=1, max_col=len(headers) + len(date_headers), max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Open a QFileDialog to select the file path
            options = QFileDialog.Options()
            excel_file, _ = QFileDialog.getSaveFileName(self.parent, "Save Excel File", "",
                                                        "Excel Files (*.xlsx);;All Files ()", options=options)

            if excel_file:  # Check if a file path was provided
                wb.save(excel_file)  # Save the file

                # Provide feedback to the user
                QMessageBox.information(self.parent, "Export Complete",
                                        f"Timecard data exported successfully to {excel_file}")
        except Exception as e:
            logging.error(f"Error in export_to_excel: {e}")
            QMessageBox.critical(self.parent, "Error", f"Failed to export to Excel: {str(e)}")

    def createTimeSheet(self):
        try:
            if self.parent.TimeListTable.rowCount() == 0:
                QMessageBox.warning(self.parent, "No rows available", "No rows detected!")
                return
            self.TimeComputation = TimeComputation(self.parent)
            self.processTimeSheetLoader = processTimeSheetLoader(self.TimeComputation, self.parent)
            self.processTimeSheetLoader.exec_()

        except Exception as e:
            logging.error(f"Unexpected error in createTimeSheet: {e}")
            QMessageBox.critical(self.parent, "Critical Error", f"An unexpected error occurred: {e}")

    def CheckSched(self, checked=False):
        try:
            selected_row = self.parent.TimeListTable.currentRow()

            if selected_row == -1:
                QMessageBox.warning(self.parent, "No Row Selected", "Please select a row from the table first!")
                return

            def get_cell_value(row, col):
                cell_widget = self.parent.TimeListTable.cellWidget(row, col)
                if isinstance(cell_widget, QComboBox):
                    return cell_widget.currentText()
                item = self.parent.TimeListTable.item(row, col)
                return item.text() if item else "N/A"

            bioNum = get_cell_value(selected_row, 1)
            empName = get_cell_value(selected_row, 2)
            trans_date = get_cell_value(selected_row, 3)
            checkIn = get_cell_value(selected_row, 5)
            checkOut = get_cell_value(selected_row, 6)

            logging.info(f"Extracted data: bioNum={bioNum}, empName={empName}, trans_date={trans_date}, "
                         f"checkIn={checkIn}, checkOut={checkOut}")

            try:
                # Calculate total_hours only
                total_hours = self.time_computation.get_total_hours_worked(checkIn, checkOut)  # Updated here
            except Exception as e:
                logging.error(f"Error calculating hours: {e}")
                QMessageBox.warning(self.parent, "Calculation Error", f"Error calculating hours: {e}")
                return

            empNum = "DefaultEmpNum"
            data = [empNum, bioNum, empName, trans_date, checkIn, checkOut, total_hours]

            if len(data) == 7:
                try:
                    dialog = schedule(data)
                    dialog.exec_()
                except Exception as e:
                    logging.error(f"Error opening schedule dialog: {e}")
                    QMessageBox.warning(self.parent, "Dialog Error", f"Failed to open the schedule dialog: {e}")
            else:
                QMessageBox.warning(self.parent, "Data Error",
                                    f"Insufficient data to process. Expected 7 items, got {len(data)}. Please check the row data.")

        except Exception as e:
            logging.error(f"Unexpected error in schedule: {e}")
            QMessageBox.critical(self.parent, "Critical Error", f"An unexpected error occurred: {e}")


class TimeComputation:
    def __init__(self, parent):
        self.parent = parent

    from datetime import datetime

    def check_holiday_type(self, trans_date):
        # Use create_connection to establish a connection
        connection = create_connection('NTP_HOLIDAY_LIST')
        if not connection:
            logging.error("Error: Unable to connect to NTP_HOLIDAY_LIST database.")
            return None  # Early exit if the connection fails

        try:
            cursor = connection.cursor()

            # Corrected format of date
            if isinstance(trans_date, str):
                parsed_date = datetime.strptime(trans_date, "%Y-%m-%d").strftime("%Y-%m-%d")
            else:
                return None

            # Direct query without parameters
            query = f"SELECT holidayName, dateType FROM type_of_dates WHERE holidayDate = '{parsed_date}'"


            cursor.execute(query)
            result = cursor.fetchone()

            if result:
                holiday_name = result[0]  # Holiday name
                date_type = result[1]  # Date type
                print(f"Holiday Name: {holiday_name}, Date Type: {date_type}")
                if date_type == 'Special Holiday':
                    return 'Special Holiday'
                elif date_type == 'Regular Holiday':
                    return 'Regular Holiday'
            return None

        except Exception as e:
            logging.error(f"Error checking holiday type for date {trans_date}: {e}")
            return None  # Return None if an error occurs

        finally:
            # Ensure the connection is closed after use
            if connection:
                connection.close()

    def calculate_overtime_hours(self, check_in_str, check_out_str):
        try:
            check_in = datetime.strptime(check_in_str, "%Y-%m-%d %H:%M:%S")
            check_out = datetime.strptime(check_out_str, "%Y-%m-%d %H:%M:%S")

            if check_out <= check_in:
                check_out += timedelta(days=1)

            total_hours = (check_out - check_in).total_seconds() / 3600

            regular_hours = 8

            overtime_hours = max(0, total_hours - regular_hours)

            return round(overtime_hours, 2)

        except Exception as e:
            logging.error(f"Error calculating overtime hours: {e}")
            return 0

    def validate_schedule(self, processor, sched_in, sched_out, check_in, check_out, bio_num, trans_date):
        """
        Validate if Sched_In and Sched_Out are within acceptable tolerance of Check_In and Check_Out.
        """
        # Tolerance for schedule validation (60 minutes)
        time_tolerance = timedelta(minutes=60)

        try:
            sched_in_time = datetime.strptime(sched_in, '%H:%M:%S')
            sched_out_time = datetime.strptime(sched_out, '%H:%M:%S')
            check_in_time = datetime.strptime(check_in, '%H:%M:%S')
            check_out_time = datetime.strptime(check_out, '%H:%M:%S')

            # Log the parsed times for debugging
            logging.debug(f"BioNum: {bio_num}, TransDate: {trans_date}, Sched_In: {sched_in_time}, "
                          f"Check_In: {check_in_time}, Sched_Out: {sched_out_time}, "
                          f"Check_Out: {check_out_time}")

            # Compare Sched_In and Check_In
            in_match = abs(sched_in_time - check_in_time) <= time_tolerance
            # out_match = abs(sched_out_time - check_out_time) <= time_tolerance

            if not in_match: # or not out_match
                # Log the mismatches for debugging
                if not in_match:
                    # logging.debug(f"Check-In mismatch: Sched_In {sched_in_time}, Check_In {check_in_time}")
                    print(f"Check-In mismatch: Sched_In {sched_in_time}, Check_In {check_in_time}")

                # if not out_match:
                #     logging.debug(f"Check-Out mismatch: Sched_Out {sched_out_time}, Check_Out {check_out_time}")
                #     print(f"Check-Out mismatch: Sched_Out {sched_out_time}, Check_Out {check_out_time}")

                processor.error.emit(f"BioNum: {bio_num}, TransDate: {trans_date} has unmatched schedule.")
                return False
        except Exception as e:
            logging.error(f"Error in schedule validation for {bio_num} on {trans_date}: {e}")
            processor.error.emit(f"Error validating schedule: {e}")
            return False

        # Schedule is valid
        return True

    def calculate_hours(self, check_in_str, check_out_str):
        # Convert strings to datetime objects
        check_in = datetime.strptime(check_in_str, "%Y-%m-%d %H:%M:%S")
        check_out = datetime.strptime(check_out_str, "%Y-%m-%d %H:%M:%S")

        # Adjust check_out for midnight crossing
        if check_out <= check_in:
            check_out += timedelta(days=1)

        # Define ND and NDOT periods
        nd_start1 = datetime.strptime("21:00", "%H:%M").time()  # 10:00 PM
        nd_end1 = datetime.strptime("23:59", "%H:%M").time()  # 11:59 PM
        nd_start2 = datetime.strptime("01:00", "%H:%M").time()  # 1:00 AM
        nd_end2 = datetime.strptime("06:00", "%H:%M").time()  # 6:00 AM
        ndot_start = datetime.strptime("02:00", "%H:%M").time()  # 2:00 AM

        total_hours = (check_out - check_in).total_seconds() / 3600
        nd_hours = 0
        ndot_hours = 0
        current_time = check_in

        while current_time < check_out:
            next_hour = min(current_time + timedelta(hours=1), check_out)

            # Extract current time for ND/NDOT calculations
            current_time_struct = current_time.time()

            # ND Hours Calculation
            if nd_start1 <= current_time_struct <= nd_end1:  # From 10 PM to midnight
                nd_period_end = min(next_hour, datetime.combine(current_time.date(), nd_end1))
                nd_hours += (nd_period_end - current_time).total_seconds() / 3600
            elif nd_start2 <= current_time_struct < nd_end2:  # From 1 AM to 6 AM
                nd_period_end = min(next_hour, datetime.combine(current_time.date(), nd_end2))
                nd_hours += (nd_period_end - current_time).total_seconds() / 3600

            # NDOT Hours Calculation
            if ndot_start <= current_time_struct < nd_end2:  # 2 AM to 6 AM
                ndot_period_end = min(next_hour, datetime.combine(current_time.date(), nd_end2))
                ndot_hours += (ndot_period_end - current_time).total_seconds() / 3600

            current_time = next_hour

        # Cap ND hours at 8
        nd_hours = min(nd_hours, 8)

        # Round and return total, ND, and NDOT hours
        return int(total_hours), int(nd_hours), int(ndot_hours)

    def calculate_late_and_undertime(self, sched_in, sched_out, check_in, check_out):
        """
        Calculate late and undertime based on scheduled times and actual times.
        """
        late = 0
        undertime = 0
        try:
            sched_in_time = datetime.strptime(sched_in, '%H:%M:%S')
            sched_out_time = datetime.strptime(sched_out, '%H:%M:%S')
            check_in_time = datetime.strptime(check_in, '%H:%M:%S')
            check_out_time = datetime.strptime(check_out, '%H:%M:%S')

            # Calculate late (if check_in is after sched_in)
            if check_in_time > sched_in_time:
                late = (check_in_time - sched_in_time).total_seconds() / 3600

            # Calculate undertime (if check_out is before sched_out)
            if check_out_time < sched_out_time:
                undertime = (sched_out_time - check_out_time).total_seconds() / 3600

        except Exception as e:
            logging.error(f"Error in calculating late/undertime: {e}")

        return late, undertime

    def get_total_hours_worked(self, check_in, check_out):
        # Parse the check-in and check-out times
        time_in = datetime.strptime(check_in, "%H:%M:%S")
        time_out = datetime.strptime(check_out, "%H:%M:%S")

        # Handle the case where the time crosses midnight
        if time_out <= time_in:
            time_out += timedelta(days=1)

        # Calculate total hours worked
        total_hours = (time_out - time_in).total_seconds() / 3600

        return round(total_hours, 3)


class searchBioNum:
    def __init__(self, parent):
        self.parent = parent
        self.populate_list_instance = populateList(self.parent)

    def search_bioNum(self):
        search_text = self.parent.searchBioNum.text().strip().lower()

        if not search_text:
            # unhides all the row if search_text is empty
            for row in range(self.parent.TimeListTable.rowCount()):
                self.parent.TimeListTable.setRowHidden(row, False)

        for row in range(self.parent.TimeListTable.rowCount()):
            item = self.parent.TimeListTable.item(row, 1)  # Bio Num column at index 1
            if item and search_text in item.text().lower():
                self.parent.TimeListTable.setRowHidden(row, False)
            else:
                self.parent.TimeListTable.setRowHidden(row, True)


class FilterDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(400, 300)
        ui_file = globalFunction.resource_path("MainFrame\\Resources\\UI\\filter.ui")
        loadUi(ui_file, self)

        self.cmbCheckIn = self.findChild(QComboBox, 'cmbCheckIn')
        self.cmbCheckOut = self.findChild(QComboBox, 'cmbCheckOut')
        self.btnOK = self.findChild(QPushButton, 'btnOK')
        self.btnClear = self.findChild(QPushButton, 'btnClear')
        self.btnMissing = self.findChild(QPushButton, 'btnMissing')

        self.btnOK.clicked.connect(self.accept)
        self.btnClear.clicked.connect(self.clear_filter)
        self.btnMissing.clicked.connect(self.show_missing)

        for combo in [self.cmbCheckIn, self.cmbCheckOut]:
            if combo.itemText(0) != "AM/PM":
                combo.insertItem(0, "AM/PM")

        self.parent = parent

    def clear_filter(self, checked=False):
        try:
            if self.parent:
                logging.info("Clearing filter...")
                logging.info(f"Original data before clearing: {self.parent.original_data}")

                # Reset to original data
                self.parent.filtered_data = self.parent.original_data.copy()

                # Populate the table using the populate_time_list_table method
                self.parent.populateComboBox.populate_time_list_table(self.parent.filtered_data)
                QMessageBox.information(self.parent, "Success", "Filter cleared and table restored to original state.")
        except Exception as e:
            QMessageBox.critical(self.parent, "Error",
                                 f"Error in clearing filter: {str(e)}. Please try again or contact the system administrator.")
            logging.error(traceback.format_exc())
        self.accept()

    def show_missing(self, checked=False):
        try:
            filter_values = {
                'check_in_ampm': "AM/PM",
                'check_out_ampm': "AM/PM",
                'show_missing': True
            }
            logging.info("Showing missing entries with filter values: %s", filter_values)
            self.apply_filter(filter_values)
        except Exception as e:
            QMessageBox.critical(self.parent, "Error",
                                 f"Error in showing missing entries: {str(e)}. "
                                 "Please try again or contact the system administrator.")
            logging.error(traceback.format_exc())
        self.accept()

    def get_filter_values(self):
        values = {
            'check_in_ampm': self.cmbCheckIn.currentText(),
            'check_out_ampm': self.cmbCheckOut.currentText(),
            'show_missing': False
        }
        logging.info(f"Filter values: {values}")
        return values

    def apply_filter(self, filter_values):
        try:
            logging.info(f"Applying filter with values: {filter_values}")

            filtered = []
            for row in self.parent.original_data:
                check_in_time = row[4]
                check_out_time = row[5]

                # Filter logic for missing entries
                if filter_values['show_missing']:
                    if check_in_time == 'Missing' or check_out_time == 'Missing':
                        filtered.append(row)
                        continue

                # AM/PM filtering logic
                if check_in_time != 'Missing' and check_out_time != 'Missing':
                    check_in_hour = int(check_in_time.split(':')[0])
                    check_out_hour = int(check_out_time.split(':')[0])

                    if filter_values['check_in_ampm'] == 'AM' and not (0 <= check_in_hour < 12):
                        continue
                    if filter_values['check_in_ampm'] == 'PM' and not (12 <= check_in_hour < 24):
                        continue
                    if filter_values['check_out_ampm'] == 'AM' and not (0 <= check_out_hour < 12):
                        continue
                    if filter_values['check_out_ampm'] == 'PM' and not (12 <= check_out_hour < 24):
                        continue

                    filtered.append(row)

            logging.info(f"Filtered data contains {len(filtered)} rows.")
            self.parent.filtered_data = filtered

            # Update the table with filtered data
            self.parent.populateComboBox.populate_table_with_data(self.parent.filtered_data)

        except Exception as e:
            logging.error(f"Error in apply_filter: {str(e)}")
            QMessageBox.critical(self.parent, "Error", f"An error occurred while applying the filter: {str(e)}")

    def filterModal(self):
        try:
            if self.exec_() == QDialog.Accepted:
                filter_values = self.get_filter_values()
                logging.info(f"Filter values obtained from dialog: {filter_values}")
                self.apply_filter(filter_values)
            else:
                logging.info("Filter dialog was canceled by the user.")

        except Exception as e:
            logging.error(f"Error in filterModal: {str(e)}")
            QMessageBox.critical(self.parent, "Error", f"An error occurred while opening the filter dialog: {str(e)}")


class TablePopulationLoader(QDialog):
    """Displays a dialog with progress bar for visualization of populating table with data"""

    def __init__(self, data, timeCardWindow=None):
        super(TablePopulationLoader, self).__init__(timeCardWindow)
        ui_file = globalFunction.resource_path("MainFrame\\Resources\\UI\\showNotification.ui")
        loadUi(ui_file, self)
        self.setFixedSize(400, 124)
        self.setModal(True)
        self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint)

        try:
            self.parent = timeCardWindow
            self.data = data

            self.functions = populateList(self.parent)

            # Get UI elements
            self.progressBar = self.findChild(QProgressBar, 'progressBar')

            self.progressBar.setVisible(True)
            self.progressBar.setValue(0)

            self.thread = QThread()
            self.worker = FetchDataToPopulateTableProcessor(self.parent, self.data)
            self.worker.moveToThread(self.thread)
            self.worker.progressChanged.connect(self.updateProgressBar)
            self.worker.finished.connect(self.fetchingDataFinished)
            self.worker.error.connect(self.fetchingDataError)
            self.thread.started.connect(self.worker.process_populate_time_list_table)
            self.thread.start()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Showing TablePopulationLoader:\n{str(e)}")
            print("Error Showing TablePopulationLoader: ", e)

    def updateProgressBar(self, value):
        self.progressBar.setValue(value)
        if value == 100:
            self.progressBar.setFormat("Finishing Up..")
        QApplication.processEvents()

    def fetchingDataFinished(self, time_data):
        self.progressBar.setVisible(False)
        self.thread.quit()
        self.thread.wait()

        # Update table
        self.parent.TimeListTable.setUpdatesEnabled(False)
        self.functions.populate_table_with_data(time_data)
        self.parent.TimeListTable.setUpdatesEnabled(True)
        self.parent.original_data = time_data

        # Closes and reset the instance of dialog
        self.close()
        self.parent.TablePopulationLoader = None

    def fetchingDataError(self, error):
        self.progressBar.setVisible(False)
        self.thread.quit()
        self.thread.wait()

        QMessageBox.critical(self.parent, "Populating Table with Data Error", error)

        # Closes and reset the instance of dialog
        self.close()
        self.parent.TablePopulationLoader = None


class FetchDataToPopulateTableProcessor(QObject):
    progressChanged = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, timeCardWindow, data):
        super().__init__()
        self.parent = timeCardWindow
        self.data = data

    def process_populate_time_list_table(self):
        """Populate the time list table with check-in and check-out times, machCode, and employee data from NTP_LOG_IMPORTS."""
        try:
            QThread.msleep(100)
            selected_year_month = self.parent.yearCC.currentText()
            from_day = self.parent.dateFromCC.currentText()
            to_day = self.parent.dateToCC.currentText()
        except Exception as e:
            print(e)

        if not selected_year_month or not from_day or not to_day:
            self.error.emit(f"selected_year_month: {selected_year_month}\nfrom_day: {from_day}\nto_day: {to_day}")
            return

        logging.info(f"Populating time list table for: {selected_year_month}, from {from_day} to {to_day}")

        from_date = f"{selected_year_month}-{from_day.zfill(2)}"
        to_date = f"{selected_year_month}-{to_day.zfill(2)}"
        table_name = f"table_{selected_year_month.replace('-', '_')}"

        # Connect to NTP_LOG_IMPORTS database
        connection_list_log = create_connection('NTP_LOG_IMPORTS')

        if not connection_list_log:
            self.error.emit("Failed to connect to database.")
            return

        try:
            cursor_list_log = connection_list_log.cursor()

            # Check if the table exists
            cursor_list_log.execute(f"SHOW TABLES LIKE '{table_name}'")
            if not cursor_list_log.fetchone():
                self.error.emit(f"Error: Table does not exist: {table_name}")
                return

            # Fetch records from the log table
            query_list_log = f"""
                SELECT ID, bioNum, Name, date, time_in, time_out, machCode, sched_in, sched_out
                FROM `{table_name}`
                WHERE date BETWEEN '{from_date}' AND '{to_date}'
                ORDER BY bioNum, date, time_in
            """
            cursor_list_log.execute(query_list_log)
            records = cursor_list_log.fetchall()

            total_records = len(records)
            if total_records == 0:
                self.finished.emit([])  # No data to process
                return

            # Prepare time data
            time_data = []

            for i, (ID, bioNum, name, trans_date, time_in, time_out, mach_code, sched_in, sched_out) in enumerate(records):
                check_in_time = time_in or "00:00:00"
                check_out_time = time_out or "00:00:00"

                # Prepare each row's data for the table
                time_data.append([
                    ID, bioNum, name, trans_date, mach_code, check_in_time, check_out_time, sched_in, sched_out
                ])

                # Update progress
                progress = int(((i + 1) / total_records) * 100)
                self.progressChanged.emit(progress)
                QThread.msleep(1)

            # Emit the finished signal with the populated time data
            self.finished.emit(time_data)

        except Exception as e:
            print(f"Error populating time list table: {e}")
            self.error.emit(f"Error populating time list table: {e}")

        finally:
            cursor_list_log.close()
            connection_list_log.close()

