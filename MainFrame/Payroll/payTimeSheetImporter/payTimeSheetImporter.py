import sys
import os

from openpyxl.workbook import Workbook

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from MainFrame.Resources.lib import *

from MainFrame.Payroll.paytimeSheet.paytimeSheet import PaytimeSheet
from MainFrame.systemFunctions import globalFunction
from MainFrame.Database_Connection.user_session import UserSession
from MainFrame.Database_Connection.DBConnection import create_connection

class FileProcessor(QObject):
    progressChanged = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, fileName):
        super().__init__()
        self.fileName = fileName

    def process(self):
        try:
            # Load the workbook and select the active sheet
            workbook = openpyxl.load_workbook(self.fileName, data_only=True)
            sheet = workbook.active

            total_rows = sheet.max_row
            content = []

            # Read the header row separately
            headers = [cell.value for cell in sheet[1]]  # Read header row
            content.append(headers)  # Add header row to content

            for row_idx in range(2, total_rows + 1):  # Skip header row
                row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                content.append(row)
                progress = int((row_idx / total_rows) * 100)
                self.progressChanged.emit(progress)
                QThread.msleep(1)  # Simulate work being done

            self.finished.emit(content)
        except Exception as e:
            self.error.emit(f"Failed to process .xlsx file: {e}")


class PayrollDialog(QDialog):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        # self.setFixedSize(418, 339)
        ui_file = globalFunction.resource_path("MainFrame\\Resources\\UI\\dialogImporter.ui")
        loadUi(ui_file, self)

        self.user_session = UserSession().getALLSessionData()
        self.user_role = str(self.user_session.get("user_role", ""))

        # Print the user_role to verify it
        print(f"User Role: {self.user_role}")

        self.configureButtons(self.user_role)

        self.importBTN.clicked.connect(self.importTxt)
        self.importBTN.setText("Import Excel")

        self.btnExportToExcel.clicked.connect(self.exportDeductionToExcel)

        self.progressBar = self.findChild(QProgressBar, 'progressBar')
        self.progressBar.setVisible(False)

    def configureButtons(self, user_role):
        """Configure button visibility based on the user role."""
        if user_role == "Pay Master 1":
            if hasattr(self, 'btnProcessTimeCard'):
                self.btnProcessTimeCard.setVisible(False)
            if hasattr(self, 'btnExportToExcel'):
                self.btnExportToExcel.setVisible(False)
        elif user_role == "Pay Master 2":
            if hasattr(self, 'btnProcessTimeCard'):
                self.btnProcessTimeCard.setVisible(False)
            if hasattr(self, 'btnExportToExcel'):
                self.btnExportToExcel.setVisible(True)

    def importTxt(self):
        fileName, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xls *.xlsx)")

        if not fileName:
            QMessageBox.information(self, "No File Selected", "Please select an Excel file to import.")
            return

        # Use FileProcessor to handle the file reading
        self.progressBar.setVisible(True)
        self.progressBar.setValue(0)
        self.thread = QThread()
        self.worker = FileProcessor(fileName)
        self.worker.moveToThread(self.thread)
        self.worker.progressChanged.connect(self.updateProgressBar)
        self.worker.finished.connect(self.fileProcessingFinished)
        self.worker.error.connect(self.fileProcessingError)
        self.thread.started.connect(self.worker.process)
        self.thread.start()

    def updateProgressBar(self, value):
        self.progressBar.setValue(value)
        if value == 100:
            self.progressBar.setFormat("Finishing Up..")
        QApplication.processEvents()

    def fileProcessingFinished(self, content):
        self.progressBar.setVisible(False)
        self.thread.quit()
        self.thread.wait()

        # Validate columns and show data
        if content:
            headers = content[0]
            data = content[1:]
            required_columns = [
                'bionum', 'empnumber', 'empname', 'costcenter', 'fromdate', 'todate', 'dayspresent', 'restday',
                'holiday', 'rsthlyday', 'orddaynite', 'rstdaynite', 'hlydaynite', 'rsthlydayn', 'orddayot',
                'rstdayot', 'hlydayot', 'rsthlydayo', 'orddaynit2', 'rstdaynit2', 'hlydaynit2', 'rsthlyday2',
                'late', 'undertime', 'absent', 'dateposted', 'remarks', 'rstshlyday', 'rstshlyda2', 'rstshlyda3',
                'rstshlyda4', 'empcompany', 'legalholid'
            ]

            # Check for None and safely strip and lower case headers
            formatted_headers = [header.strip().lower() if header is not None else '' for header in headers]

            missing_columns = [col for col in required_columns if col not in formatted_headers]

            if missing_columns:
                error_message = f"Missing required columns: {', '.join(missing_columns)}"
                QMessageBox.warning(self, "Missing Columns", error_message)
                return

            self.showData(content, self.user_role)

    def fileProcessingError(self, error):
        logging.error(f"Failed to read file: {error}")
        self.progressBar.setVisible(False)
        self.thread.quit()
        self.thread.wait()

        QMessageBox.critical(self, "File Processing Error", f"An error occurred while processing the file:\n{error}")
        self.close()

    def showData(self, content, user_role):
        self.paytimesheet = PaytimeSheet(self.main_window, content, user_role)  # Pass user_role
        self.main_window.open_dialogs.append(self.paytimesheet)
        print(content)
        self.paytimesheet.show()
        self.close()

    def exportDeductionToExcel(self):
        connection = create_connection('SYSTEM_STORE_DEDUCTION')
        if connection is None:
            print("Failed to connect to SYSTEM_STORE_DEDUCTION database.")
            return

        cursor = connection.cursor()

        try:
            query = "SELECT * FROM deductions"
            cursor.execute(query)
            result = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]

            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Deductions"

            # Write column headers
            ws.append(column_names)

            # Write data rows
            for row in result:
                ws.append(row)

            # Prompt the user to choose a file location and name
            file_name, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx);;All Files (*)")
            if file_name:
                # Save the workbook
                wb.save(file_name)
                QMessageBox.information(self, "Export Successful", "Deductions Data was exported successfully!")
                print(f"Data exported successfully to {file_name}")

        except Error as e:
            print(f"Error fetching or processing deduction data: {e}")

        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()