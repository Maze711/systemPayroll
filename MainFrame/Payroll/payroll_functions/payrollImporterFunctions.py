from MainFrame.Resources.lib import *
from openpyxl.workbook import Workbook
from MainFrame.Payroll.paytimeSheet.paytimeSheet import PaytimeSheet
from MainFrame.Database_Connection.DBConnection import create_connection

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class FileProcessor(QObject):
    progressChanged = pyqtSignal(int)
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, fileName):
        super().__init__()
        self.fileName = fileName
        # When using a book1 excel file the daypresent should be dayspresent
        self.column_mapping = {
            'Bio_No.': ['Bio_No.'],
            'Emp_Number': ['Emp_Number'],
            'Emp_Name': ['Emp_Name'],
            'Cost_Center': ['Cost_Center'],
            'Days_Work': ['Days_Work'],
            'Days_Present': ['Days_Present'],
            'Hours_Work': ['Hours_Work'],
            'Late': ['Late'],
            'Undertime': ['Undertime'],
            'OrdDay_Hrs': ['OrdDay_Hrs'],
            'OrdDayOT_Hrs': ['OrdDayOT_Hrs'],
            'OrdDayND_Hrs': ['OrdDayND_Hrs'],
            'OrdDayNDOT_Hrs': ['OrdDayNDOT_Hrs'],
            'RstDay_Hrs': ['RstDay_Hrs'],
            'RstDayOT_Hrs': ['RstDayOT_Hrs'],
            'RstDayND_Hrs': ['RstDayND_Hrs'],
            'RstDayNDOT_Hrs': ['RstDayNDOT_Hrs'],
            'SplHlyday_Hrs': ['SplHlyday_Hrs'],
            'SplHlydayOT_Hrs': ['SplHlydayOT_Hrs'],
            'SplHlydayND_Hrs': ['SplHlydayND_Hrs'],
            'SplHlydayNDOT_Hrs': ['SplHlydayNDOT_Hrs'],
            'RegHlyday_Hrs': ['RegHlyday_Hrs'],
            'RegHlydayOT_Hrs': ['RegHlydayOT_Hrs'],
            'RegHlydayND_Hrs': ['RegHlydayND_Hrs'],
            'RegHlydayNDOT_Hrs': ['RegHlydayNDOT_Hrs'],
            'SplHldyRD_Hrs': ['SplHldyRD_Hrs'],
            'SplHldyRDOT_Hrs': ['SplHldyRDOT_Hrs'],
            'SplHldyRDND_Hrs': ['SplHldyRDND_Hrs'],
            'SplHldyRDNDOT_Hrs': ['SplHldyRDNDOT_Hrs'],
            'RegHldyRD_Hrs': ['RegHldyRD_Hrs'],
            'RegHldyRDOT_Hrs': ['RegHldyRDOT_Hrs'],
            'RegHldyRDND_Hrs': ['RegHldyRDND_Hrs'],
            'RegHldyRDNDOT_Hrs': ['RegHldyRDNDOT_Hrs'],
        }

        self.required_columns = list(self.column_mapping.keys())

    def process(self):
        try:
            content = []
            file_ext = self.fileName.lower()

            if file_ext.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(self.fileName, data_only=True)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                total_rows = sheet.max_row
                for row_idx in range(2, total_rows + 1):  # Skip header row
                    row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                    content.append(row)
                    progress = int((row_idx / total_rows) * 100)
                    self.progressChanged.emit(progress)
                    QThread.msleep(1)  # Simulate work being done
            elif file_ext.endswith('.xls'):
                workbook = xlrd.open_workbook(self.fileName, encoding_override="cp1252")
                sheet = workbook.sheet_by_index(0)
                headers = sheet.row_values(0)
                total_rows = sheet.nrows
                for row_idx in range(1, total_rows):  # Skip header row
                    row = sheet.row_values(row_idx)
                    content.append(row)
                    progress = int((row_idx / total_rows) * 100)
                    self.progressChanged.emit(progress)
                    QThread.msleep(1)  # Simulate work being done
            else:
                raise ValueError(f"Unsupported file format: {file_ext.split('.')[-1]}")

            # Validate and standardize columns
            standardized_headers = self.standardize_headers(headers)
            missing_columns = [col for col in self.required_columns if col not in standardized_headers]

            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

            # Replace original headers with standardized headers
            content[0] = standardized_headers
            self.finished.emit(content)
        except ValueError as ve:
            self.error.emit(str(ve))
        except Exception as e:
            self.error.emit(f"Unexpected error: {e}")

    def standardize_headers(self, headers):
        standardized = []
        for header in headers:
            header_lower = header.strip().lower() if header else ''
            for std_name, variations in self.column_mapping.items():
                if header_lower in variations:
                    standardized.append(std_name)
                    break
            else:
                standardized.append(header_lower)  # Keep original if no match found
        return standardized


class PayrollImporterFunctions:
    def __init__(self, dialog, user_role):
        self.dialog = dialog
        self.user_role = user_role

    def importTxt(self):
        fileName, _ = QFileDialog.getOpenFileName(self.dialog, "Select Excel File", "", "Excel Files (*.xls *.xlsx)")

        if not fileName:
            QMessageBox.information(self.dialog, "No File Selected", "Please select an Excel file to import.")
            return

        # Use FileProcessor to handle the file reading
        self.dialog.progressBar.setVisible(True)
        self.dialog.progressBar.setValue(0)
        self.thread = QThread()
        self.worker = FileProcessor(fileName)
        self.worker.moveToThread(self.thread)
        self.worker.progressChanged.connect(self.updateProgressBar)
        self.worker.finished.connect(self.fileProcessingFinished)
        self.worker.error.connect(self.fileProcessingError)
        self.thread.started.connect(self.worker.process)
        self.thread.start()

    def updateProgressBar(self, value):
        self.dialog.progressBar.setValue(value)
        if value == 100:
            self.dialog.progressBar.setFormat("Finishing Up..")
        QApplication.processEvents()

    def fileProcessingFinished(self, content):
        self.dialog.progressBar.setVisible(False)
        self.thread.quit()
        self.thread.wait()

        # Validate columns and show data
        if content:
            headers = content[0]
            data = content[1:]
            # required_columns = [
            #     'bionum', 'empnumber', 'empname', 'costcenter', 'fromdate', 'todate', 'dayspresent', 'restday',
            #     'holiday', 'rsthlyday', 'orddaynite', 'rstdaynite', 'hlydaynite', 'rsthlydayn', 'orddayot',
            #     'rstdayot', 'hlydayot', 'rsthlydayo', 'orddaynit2', 'rstdaynit2', 'hlydaynit2', 'rsthlyday2',
            #     'late', 'undertime', 'absent', 'dateposted', 'remarks', 'rstshlyday', 'rstshlyda2', 'rstshlyda3',
            #     'rstshlyda4', 'empcompany', 'legalholid'
            # ]

            #When using a XLS this will be the column
            required_columns = ['Bio_No.', 'Emp_Number', 'Emp_Name', 'Cost_Center', 'Days_Work', 'Days_Present',
                                'Hours_Work', 'Late', 'Undertime', 'OrdDay_Hrs', 'OrdDayOT_Hrs', 'OrdDayND_Hrs',
                                'OrdDayNDOT_Hrs', 'RstDay_Hrs', 'RstDayOT_Hrs', 'RstDayND_Hrs', 'RstDayNDOT_Hrs',
                                'SplHlyday_Hrs', 'SplHlydayOT_Hrs', 'SplHlydayND_Hrs', 'SplHlydayNDOT_Hrs',
                                'RegHlyday_Hrs', 'RegHlydayOT_Hrs', 'RegHlydayND_Hrs', 'RegHlydayNDOT_Hrs',
                                'SplHldyRD_Hrs', 'SplHldyRDOT_Hrs', 'SplHldyRDND_Hrs', 'SplHldyRDNDOT_Hrs',
                                'RegHldyRD_Hrs', 'RegHldyRDOT_Hrs', 'RegHldyRDND_Hrs', 'RegHldyRDNDOT_Hrs']

            # Check for None and safely strip and lower case headers
            formatted_headers = [header.strip().lower() if header is not None else '' for header in headers]

            missing_columns = [col for col in required_columns if col not in formatted_headers]

            if missing_columns:
                error_message = f"Missing required columns: {', '.join(missing_columns)}"
                QMessageBox.warning(self, "Missing Columns", error_message)
                return

            self.showData(content, self.user_role)

    def fileProcessingError(self, error):
        logging.error(f"Failed to read file: {error}")
        self.dialog.progressBar.setVisible(False)
        self.thread.quit()
        self.thread.wait()

        QMessageBox.critical(self.dialog, "File Processing Error", f"An error occurred while processing the file:\n{error}")
        self.dialog.close()

    def showData(self, content, user_role):
        self.paytimesheet = PaytimeSheet(self.dialog.main_window, content, user_role)  # Pass user_role
        self.dialog.main_window.open_dialogs.append(self.paytimesheet)
        print(content)
        self.paytimesheet.show()
        self.dialog.close()